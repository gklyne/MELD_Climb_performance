# !/usr/bin/env python

""""
Grid class and implementations for CSV and Excel spreadsheet files

See: README.md in this directory
"""

__author__      = "Graham Klyne (GK@ACM.ORG)"
__copyright__   = "Copyright 2017, Graham Klyne"
__license__     = "MIT (http://opensource.org/licenses/MIT)"

import urlparse
import csv
import openpyxl
import logging
import traceback

log = logging.getLogger(__name__)
# log.setLevel(logging.INFO)

class Grid(object):
    """
    Interface for grid or spreadsheet object.
    """

    def __init__(self, baseuri=None):
        ### self._baseuri = ro_uriutils.resolveFileAsUri(baseuri or "")
        self._baseuri = (baseuri or "")
        return

    def baseUri(self, uriref=None):
        if uriref:
            self._baseuri = self.resolveUri(uriref)
        return self._baseuri

    def resolveUri(self, uriref):
        # return ro_uriutils.resolveUri(uriref, self._baseuri)
        endswithhash = uriref.endswith("#")
        resolveduri  = urlparse.urljoin(self._baseuri, uriref)
        if endswithhash and not resolveduri.endswith("#"):
            resolveduri = resolveduri + "#"
        return resolveduri

    def cell(self, row, col):
        assert False, "Unimplemented 'cell' method"

    def __getitem__(self, row):
        return GridRow(self, row)

    def __iter__(self):
        row = 0
        while True:
            log.info("__iter__ row %d"%(row))
            try:
                rowdata = self[row]
                _probe = rowdata[0]
                yield rowdata
                row += 1
            except ValueError, e:
                break
        return

    def rows(self, rowfrom, rowto=100000):
        for i in range(rowfrom, rowto):
            # log.info("rows(%d, %d) row %d"%(rowfrom, rowto, i))
            try:
                rowdata = self[i]
                _probe = rowdata[0]
                yield rowdata
            except Exception, e:
                log.info("Grid.rows: %s"%(e), exc_info=True)
                # log.info("".join(traceback.format_stack()))
                break
        return

class GridRow(object):
    """
    Interface for auxiliary grid or spreadsheet row.
    """

    def __init__(self, grid, row):
        self._grid = grid
        self._row  = row
        return

    def __getitem__(self, col):
        return self._grid.cell(self._row, col)

    def __iter__(self):
        col = 0
        while True:
            try:
                yield self._grid.cell(self._row, col)
                col += 1
            except ValueError, e:
                break
            except IndexError, e:
                break
        return

    def __repr__(self):
        return "grid.GridRow(" + ",".join([str(c) if c else "" for c in self]) + ")"

class GridCSV(Grid):
    """
    Initialize a grid object based on a CSV file

    @param csvfilename: Name of a file that contains CSV data
    @param baseuri:     A string used as the base URI for references in the grid.
    @param dialect:     An optional dialect parameter (e.g. 'excel', 'excel-tab').
                        If not specified, the system sniffs the content of the CSV 
                        to guess the CSV dialect used.
    """

    def __init__(self, csvfilename, baseuri=None, dialect=None):
        csvfile = open(csvfilename)
        super(GridCSV, self).__init__(baseuri=baseuri)
        if not dialect:
            dialect = csv.Sniffer().sniff(csvfile.read(1024))
            csvfile.seek(0)
        log.info("GridCSV: %s, %s"%(csvfile, dialect))
        reader = csv.reader(csvfile, dialect)
        self._rows   = []
        self._maxcol = 0
        self._maxrow = 0
        for row in reader:
            log.info("- row: %s"%(repr(row)))
            self._rows.append(row)
            if len(row) > self._maxcol: self._maxcol = len(row)
            self._maxrow += 1
        log.info("GridCSV: maxrow %d, maxcol %d"%(self._maxrow, self._maxcol))
        return

    def cell(self, row, col):
        if row < 10:
            log.info("GridCSV cell %d %d"%(row, col))
        if col >= self._maxcol:
            raise ValueError("Column out of range")
        return self._rows[row][col] if col < len(self._rows[row]) else ""

class GridTSV(Grid):
    """
    Initialize a grid object based on a TSV file

    @param tsvfilename: Name of a file that contains tab-separated data
    @param baseuri:     A string used as the base URI for references in the grid.
    @param dialect:     An optional dialect parameter (e.g. 'excel', 'excel-tab').
                        If not specified, the system sniffs the content of the TSV 
                        to guess the TSV dialect used.
    """

    def __init__(self, tsvfilename, baseuri=None, dialect=None):
        tsvfile = open(tsvfilename)
        super(GridTSV, self).__init__(baseuri=baseuri)
        if not dialect:
            dialect = csv.Sniffer().sniff(tsvfile.read(1024))
            tsvfile.seek(0)
        log.debug("GridTSV: %s, %s"%(tsvfile, dialect))
        reader = csv.reader(tsvfile, dialect)
        self._rows   = []
        self._maxcol = 0
        for row in reader:
            # log.debug("- row: %s"%(repr(row)))
            self._rows.append(row)
            if len(row) > self._maxcol: self._maxcol = len(row)
        return

    def cell(self, row, col):
        # log.info("GridTSV cell %d %d"%(row, col))
        if col > self._maxcol:
            raise ValueError("Column out of range")
        return self._rows[row][col] if col < len(self._rows[row]) else ""


class GridExcel(Grid):
    """
    Initialize a grid object based on an excel file

    @param xlsfile:     Filename of an Excel spreadsheet file
    @param baseuri:     A string used as the base URI for references in the grid.
    """

    def __init__(self, xlsfilename, baseuri=None):
        super(GridExcel, self).__init__(baseuri=baseuri)
        log.debug("GridExcel: %s"%(xlsfilename))
        self._workbook = openpyxl.load_workbook(filename=xlsfilename)
        # Assume first and only worksheet
        self._sheet  = self._workbook.active
        self._maxrow = self._sheet.max_row
        self._maxcol = self._sheet.max_column
        log.info("GridExcel sheet size: %d, %d"%(self._maxrow, self._maxcol))
        return

    def cell(self, row, col):
        if row in range(self._maxrow) and col in range(self._maxcol):
            cell = self._sheet.cell(row=row+1, column=col+1)
            # log.debug("GridExcel.cell [%d,%d] = %r:%s"%(row, col, cell.value, cell.data_type))
            # log.debug("GridExcel.cell data_type: %s"%(cell.data_type))
            if cell.data_type == openpyxl.cell.Cell.TYPE_NULL: # or cell.value is None:
                return None
            if cell.data_type == openpyxl.cell.Cell.TYPE_STRING:
                return cell.value
            if cell.data_type == openpyxl.cell.Cell.TYPE_NUMERIC:
                return str(cell.value)
            if cell.data_type == openpyxl.cell.Cell.TYPE_BOOL:
                return "True" if cell.value else "False"
            return "????%s"%(cell.data_type)
            #     raise ValueError("Cell type must be empty or string (got %d)"%(cell.data_type))
        raise IndexError("Index outside bound of spreadsheet: %d,%d (%d,%d)"%(row, col, self._maxrow, self._maxcol))

